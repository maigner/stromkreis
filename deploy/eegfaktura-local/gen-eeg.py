#!/usr/bin/env python3
"""Erzeugt aus einer EEG-Konfiguration (eegs/<RC>.json, siehe gen-members.py) die beiden
Importdateien fuer die EEG-Faktura-Testinstanz:

  <ausgabeordner>/<RC>-Stammdaten.xlsx    Sheet "EEG Stammdaten" (Backend /api/eeg/import/masterdata)
  <ausgabeordner>/<RC>-Energiedaten.xlsx  Sheet "Energiedaten"   (energystore singleUpload)

Aufruf:  gen-eeg.py <config.json> <von YYYY-MM-DD> <bis YYYY-MM-DD> <ausgabeordner>
Braucht openpyxl. Deterministisch (Seed aus der RC-Nummer).

Energiemodell: Haushalts-, Waermepumpen- und Gewerbeprofile mit Tages-/Wochen-/Jahresgang,
PV nach Sonnenstand (Salzkammergut, 47.7 N) mit Tageswetter (AR(1)-Bewoelkung),
dynamische Zuteilung proportional zum Verbrauch. Identitaeten je Viertelstunde:
Summe Anteil = Gesamte Erzeugung; Eigendeckung_i = min(Verbrauch_i, Anteil_i);
Summe Ueberschuss = Erzeugung - Summe Eigendeckung. Zeitstempel sind Ortszeit
(Europe/Vienna) als naive Excel-Datumswerte, 96 Slots je Tag (wie der EDA-Report).
Das Energiedaten-Format ist aus dem energystore-Quellcode (excel/ExcelSourceNew.go)
abgeleitet, das Stammdaten-Format aus der Upstream-Vorlage TE100200-Muster-Stammdatenimport.xlsx.
"""
import json
import math
import random
import sys
from datetime import date, datetime, timedelta

from openpyxl import Workbook

LAT = 47.71
SLOT_H = 0.25

HDR_CON = [
    "Gesamtverbrauch lt. Messung (bei Teilnahme gem. Erzeugung) [KWH]",
    "Anteil gemeinschaftliche Erzeugung [KWH]",
    "Eigendeckung gemeinschaftliche Erzeugung [KWH]",
]
HDR_GEN = [
    "Gesamte gemeinschaftliche Erzeugung [KWH]",
    "Gesamt/Überschusserzeugung, Gemeinschaftsüberschuss [KWH]",
]
DISCLAIMER = (
    "Die Informationen in den Reports werden auf Basis der Datensätze generiert, welche entsprechend den "
    "gesetzlichen Bestimmungen für intelligente Messgeräte vom Netzbetreiber übermittelt werden. "
    "TESTDATEN: synthetisch erzeugt fuer die Stromkreis-Testinstanz (deploy/eegfaktura-local/gen-eeg.py)."
)
LEER = "[### Leerzeile für Importer ###]"
MASTER_HDR = ["Netzbetreiber", "Gemeinschafts-ID", "Ortsgebiet", "PLZ", "Ort", "Straße", "Hausnummer", "Stiege",
              "Stock", "Tür", "Adresszusatz", "Zählpunkt", "Energierichtung", "EquipmentNr", "ObjektName",
              "Überschusseinspeisung", "Energiequelle", "Verteilungsmodell", "Zugeteilte Menge in Prozent",
              "TitelVor", "Name 1", "Name 2", "TitelNach", "BusinessRole", "IBAN", "Kontoinhaber", "email",
              "SteuerNr", "MitgliedsNr", "Zählpunktstatus", "Meter Codes", "registriert seit"]
CODES_CON = "1-1:1.9.0 G.01,1-1:2.9.0 G.02,1-1:2.9.0 G.03"
CODES_GEN = "1-1:2.9.0 G.01,1-1:2.9.0 P.01"


def solar_factor(day, hour, azimuth_deg):
    """Relative Einstrahlung auf die Modulflaeche (0..1) bei klarem Himmel."""
    doy = day.timetuple().tm_yday
    decl = math.radians(23.44) * math.sin(2 * math.pi * (284 + doy) / 365)
    lat = math.radians(LAT)
    dst_start = date(day.year, 3, 31) - timedelta(days=(date(day.year, 3, 31).weekday() + 1) % 7)
    dst_end = date(day.year, 10, 31) - timedelta(days=(date(day.year, 10, 31).weekday() + 1) % 7)
    dst = 1 if dst_start <= day < dst_end else 0
    solar_h = hour - dst - (15 - 13.6) / 15
    ha = math.radians(15 * (solar_h - 12))
    sin_el = math.sin(lat) * math.sin(decl) + math.cos(lat) * math.cos(decl) * math.cos(ha)
    if sin_el <= 0:
        return 0.0
    el = math.asin(sin_el)
    az = math.atan2(math.sin(ha), math.cos(ha) * math.sin(lat) - math.tan(decl) * math.cos(lat)) + math.pi
    tilt = math.radians(30)
    cos_inc = math.sin(el) * math.cos(tilt) + math.cos(el) * math.sin(tilt) * math.cos(az - math.radians(azimuth_deg))
    airmass = 1 / max(sin_el, 0.05)
    clear = 0.75 ** (airmass ** 0.678)
    return max(cos_inc, 0.0) * clear


def daily_weather(days, rng):
    """Bewoelkungsfaktor je Tag (0.08 .. 1.0), AR(1) mit Jahresgang (Winter truebe)."""
    out, x = {}, 0.0
    for d in days:
        season = 0.72 + 0.20 * math.cos(2 * math.pi * (d.timetuple().tm_yday - 190) / 365)
        x = 0.7 * x + rng.gauss(0, 0.28)
        out[d] = min(1.0, max(0.08, season + x))
    return out


def consumption_profile(kind, d, hour, rng):
    """Relativer Lastwert (Mittel ~1 ueber das Jahr)."""
    winter = math.cos(2 * math.pi * (d.timetuple().tm_yday - 15) / 365)
    weekend = d.weekday() >= 5
    if kind == "gewerbe":
        base = 0.35
        if not weekend and 7 <= hour < 18:
            base = 1.6 + 0.3 * math.sin(math.pi * (hour - 7) / 11)
        elif weekend and 9 <= hour < 13:
            base = 0.7
        base *= 1 + 0.15 * winter
    else:
        base = 0.45
        if 6 <= hour < 9:
            base = 1.3
        elif 11 <= hour < 14:
            base = 1.0 if weekend else 0.7
        elif 17 <= hour < 22:
            base = 1.9
        elif hour >= 22 or hour < 6:
            base = 0.4
        base *= 1 + 0.12 * winter
        if kind == "waermepumpe":
            base += max(0.0, winter) * (1.6 + 0.4 * (1 if hour < 7 or hour >= 16 else 0))
    return base * rng.uniform(0.75, 1.25)


def parse_begin(v):
    return date.fromisoformat(v) if v else None


def write_masterdata(cfg, out, start):
    wb = Workbook()
    ws = wb.active
    ws.title = "EEG Stammdaten"
    for _ in range(6):
        ws.append([LEER])
    ws.append(MASTER_HDR)
    ws.append([LEER]); ws.append([LEER])
    reg = start.strftime("%-d.%-m.%Y")
    for i, m in enumerate(cfg["members"], 1):
        points = [("CONSUMPTION", m["consumption"]["zp"], CODES_CON)] if m.get("consumption") else []
        if m.get("generation"):
            points.append(("GENERATION", m["generation"]["zp"], CODES_GEN))
        name1, name2 = m["first"], m["last"]
        begin = parse_begin((m.get("consumption") or m.get("generation") or {}).get("begin"))
        since = (begin or start).strftime("%-d.%-m.%Y")
        for k, (direction, zp, codes) in enumerate(points):
            ws.append([cfg["grid"], cfg["community_id"], "REGIONAL", cfg["plz"], cfg["ort"], m["street"], m["nr"],
                       None, None, None, None, zp, direction, f"ZP{i:03d}{k}", None, "POOLED", "SONNE", "DYNAMIC",
                       None, None, name1, name2, None, "business" if m.get("business") else "privat",
                       f"AT61{i:016d}" if k == 0 else None, f"{name1} {name2}",
                       f"{name1}.{name2}@example.com".lower().replace(" ", "-"),
                       None, f"{i:03d}", "ACTIVATED", codes, since if since else reg])
    h = wb.create_sheet("Historie")
    h.append(["PLZ", "Ort", "Straße", "Hausnummer", "Stiege", "Stock", "Tür", "Adresszusatz", "Zählpunkt",
              "Zeitpunkt", "Benutzer", "Prozess", "Änderungen"])
    wb.save(out)


def write_energy(cfg, out, start, end, rng):
    consumers = [(m["consumption"]["zp"], f"{m['first']} {m['last']}", m["consumption"]["annual_kwh"],
                  m["consumption"]["profile"], parse_begin(m["consumption"].get("begin")))
                 for m in cfg["members"] if m.get("consumption")]
    producers = [(m["generation"]["zp"], f"{m['first']} {m['last']}", m["generation"]["kwp"],
                  m["generation"]["azimuth"], parse_begin(m["generation"].get("begin")))
                 for m in cfg["members"] if m.get("generation")]
    days = [start + timedelta(n) for n in range((end - start).days + 1)]
    weather = daily_weather(days, rng)
    con_scale = [c[2] / (365 * 96) for c in consumers]
    gen_scale = [p[2] * 0.85 * SLOT_H for p in producers]

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Energiedaten")
    ids, names, dirs, starts, ends, codes = (["MeteringpointID"], ["Name"], ["Energy direction"],
                                             ["Period start"], ["Period end"], ["Metercode"])
    pstart = datetime(start.year, start.month, start.day)
    pend = datetime(end.year, end.month, end.day) + timedelta(days=1)

    def add(zp, name, direction, begin, hdrs):
        for h in hdrs:
            ids.append(zp); names.append(name); dirs.append(direction); codes.append(h)
            starts.append(datetime.combine(max(begin, start), datetime.min.time()) if begin else pstart)
            ends.append(pend)

    for zp, name, _, _, begin in consumers:
        add(zp, name, "CONSUMPTION", begin, HDR_CON)
    for zp, name, _, _, begin in producers:
        add(zp, name, "GENERATION", begin, HDR_GEN)
    ncols = len(ids) - 1
    ws.append([DISCLAIMER])
    ws.append(ids); ws.append(names); ws.append(dirs); ws.append(starts); ws.append(ends)
    ws.append(["MeteringReason"] + ["SM Datenübermittlung"] * ncols)
    ws.append(["Metering Interval"] + ["QH (viertelstündlich)"] * ncols)
    ws.append(["Number of Metering Intervals"] + [None] * ncols)
    ws.append(codes)
    ws.append(["Spaltensumme"] + [None] * ncols)

    tot = {"con": 0.0, "gen": 0.0, "cover": 0.0}
    for d in days:
        cloud = weather[d]
        for slot in range(96):
            hour = slot * SLOT_H
            ts = datetime(d.year, d.month, d.day) + timedelta(hours=hour)
            cons = [0.0 if (c[4] and d < c[4]) else round(s * consumption_profile(c[3], d, hour, rng), 3)
                    for c, s in zip(consumers, con_scale)]
            gens = [0.0 if (p[4] and d < p[4]) else
                    round(s * solar_factor(d, hour + 0.125, p[3]) * min(1.0, cloud * rng.uniform(0.85, 1.1)), 3)
                    for p, s in zip(producers, gen_scale)]
            C, G = sum(cons), sum(gens)
            shares = [G * c / C if C > 0 else 0.0 for c in cons]
            covers = [min(c, s) for c, s in zip(cons, shares)]
            cover_sum = sum(covers)
            surplus_ratio = (G - cover_sum) / G if G > 0 else 0.0
            row = [ts]
            for c, s, v in zip(cons, shares, covers):
                row += [c, round(s, 6), round(v, 6)]
            for g in gens:
                row += [g, round(g * surplus_ratio, 6)]
            ws.append(row)
            tot["con"] += C; tot["gen"] += G; tot["cover"] += cover_sum
    wb.save(out)
    return len(days), ncols, tot


def main():
    cfg = json.load(open(sys.argv[1]))
    start, end, outdir = date.fromisoformat(sys.argv[2]), date.fromisoformat(sys.argv[3]), sys.argv[4]
    rc = cfg["rc"]
    rng = random.Random(int("".join(ch for ch in rc if ch.isdigit())))
    m_out, e_out = f"{outdir}/{rc}-Stammdaten.xlsx", f"{outdir}/{rc}-Energiedaten.xlsx"
    write_masterdata(cfg, m_out, start)
    ndays, ncols, tot = write_energy(cfg, e_out, start, end, rng)
    npv = sum(1 for m in cfg["members"] if m.get("generation"))
    print(f"{rc} ({cfg['name']}, {cfg['ort']}): {len(cfg['members'])} Mitglieder, {npv} PV-Anlagen -> "
          f"{m_out}, {e_out}: {ndays} Tage, {ndays * 96} Zeilen, {ncols} Spalten; "
          f"Verbrauch {tot['con']:.0f} kWh, Erzeugung {tot['gen']:.0f} kWh, Eigendeckung {tot['cover']:.0f} kWh "
          f"({100 * tot['cover'] / max(tot['gen'], 1):.0f}% der Erzeugung, {100 * tot['cover'] / max(tot['con'], 1):.0f}% des Verbrauchs)")


if __name__ == "__main__":
    main()
